import streamlit as st
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import openpyxl
import re

def scrape_links(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        r = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(r.text, 'html.parser')
        links = []
        for a in soup.find_all('a', href=True):
            full = urljoin(url, a['href'])
            text = a.get_text().strip()
            if full.startswith('http') and re.search(r'AI|tool|app|platform', text+full, re.I):
                links.append((full, text))
        return list(set(links))
    except:
        return []

st.title("TechSol Scraper")
url = st.text_input("Enter Website URL")
if st.button("Scrape & Add"):
    links = scrape_links(url)
    st.write(f"Found {len(links)} links")
    if links:
        wb = openpyxl.load_workbook("TechSol_Directory_v1_0.xlsx")
        ws = wb["Directory"]
        next_row = ws.max_row + 1
        for i, (u, t) in enumerate(links):
            ws.cell(next_row+i, 4, u)
            ws.cell(next_row+i, 1, f"TSD-{str(next_row+i).zfill(6)}")
            ws.cell(next_row+i, 12, f"Scraped: {t}")
        wb.save("TechSol_Directory_v1_0.xlsx")
        st.success("Added to Directory!")